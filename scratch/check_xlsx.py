import openpyxl

def check():
    try:
        wb = openpyxl.load_workbook("plagas chiapas.xlsx")
        print("Sheets in workbook:")
        for name in wb.sheetnames:
            print("Sheet:", name)
            sheet = wb[name]
            print("Dimensions:", sheet.dimensions)
            for r in range(1, min(10, sheet.max_row + 1)):
                row_vals = [sheet.cell(r, c).value for c in range(1, min(10, sheet.max_column + 1))]
                print(f"Row {r}:", row_vals)
    except Exception as e:
        print("Error checking xlsx:", e)

if __name__ == "__main__":
    check()
